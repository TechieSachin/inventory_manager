import openpyxl
import glob
import tabula # for pdf to excel

pdf_name="INV2309194-INS-5595.pdf"
xlsxname="agile_rcvd.xlsx"
input_agile_path = "input/excel/agile/"
output_excel_path = "output/excel/"

def pdf_to_excel(pdf_name,xlsxname):
    # Read PDF File
    df = tabula.read_pdf(input_agile_path+pdf_name, pages = 1)[0]

    # Convert into Excel File
    df.to_excel(input_agile_path+xlsxname)

pdf_to_excel(pdf_name,xlsxname)

# excel files in the path
file_list = glob.glob(input_agile_path + "/*.xlsx")
print(file_list)

wb_agile_sent = openpyxl.load_workbook(file_list[0])
#wb_agile_sent = openpyxl.load_workbook('input/excel/AgileOrderList.xlsx')
ws_agile_sent = wb_agile_sent.active
# print('----------------------------------------------------------------')

wb_agile_rcvd = openpyxl.load_workbook(file_list[1])
#wb_agileorder_rcvd = openpyxl.load_workbook('input/excel/agile/INV2308142-INS-5595.xlsx')
ws_agile_rcvd = wb_agile_rcvd.active

wb_agile_notrcvd = openpyxl.Workbook()
ws_agile_notrcvd = wb_agile_notrcvd.active

flag = False
# Input:  invoice_wacl_sent.xlsx: Order placed with Agile
# Input:  Agile_rcvd.xlsx: Order recived from Agile
# Output: Agile_notrcvd.xlsx: Order NOT received from Agile
for sent_row in ws_agile_sent.iter_rows(min_row=3,max_row=ws_agile_sent.max_row+1,min_col=1,max_col=ws_agile_sent.max_column+1,values_only=True): 

    item_sent = sent_row[4]
    if item_sent == 80: #Skip fuel
        continue
    flag = False

    for rcvd_row in ws_agile_rcvd.iter_rows(min_row=2,max_row=ws_agile_rcvd.max_row+1,min_col=1,max_col=ws_agile_rcvd.max_column+1,values_only=True):

        item_rcvd = rcvd_row[6]
#        print("agileorder_rcvd item:" + str(item_rcvd))
        if item_sent == item_rcvd:
            flag = True
            break
    if flag == False:
        ws_agile_notrcvd.append(sent_row)
            
wb_agile_notrcvd.save(output_excel_path + 'Agile_notrcvd.xlsx')

'''
# Input: invoice_wacl_sent.xlsx': Order placed with Wacl
# Output: Wacl_rcvd.xlsx: Order recived from Wacl
# Output: Wacl_notrcvd.xlsx: Order NOT received from Wacl
wb_wacl_sent = openpyxl.load_workbook('input/excel/wacl/invoice_wacl_sent.xlsx')
ws_wacl_sent = wb_wacl_sent.active
wb_wacl_rcvd = openpyxl.Workbook()
ws_wacl_rcvd = wb_wacl_rcvd.active
wb_wacl_notrcvd = openpyxl.Workbook()
ws_wacl_notrcvd = wb_wacl_notrcvd.active

for sent_row in ws_wacl_sent.iter_rows(min_row=3,max_row=ws_wacl_sent.max_row+1,min_col=1,max_col=ws_wacl_sent.max_column+1,values_only=True):
#        print(rcvd_row)

    qty_invoiced = sent_row[8]
#    print("qty_invoiced: " + str(qty_invoiced))
    if str(qty_invoiced) == "QtyInvoiced":
        ws_wacl_rcvd.append(sent_row)
        ws_wacl_notrcvd.append(sent_row)
        continue

    if qty_invoiced == 0:
        ws_wacl_notrcvd.append(sent_row)
    else:
        ws_wacl_rcvd.append(sent_row)

wb_wacl_rcvd.save(output_excel_path + 'Wacl_rcvd.xlsx')
wb_wacl_notrcvd.save(output_excel_path+'Wacl_notrcvd.xlsx')

# Input:  Agile_notrcvd.xlsx: Order NOT received from Agile
# Input:  Wacl_rcvd.xlsx: Order recived from Wacl
# Output: Order_notrcvd_1.xlsx: Order NOT received from  Agile and Wacl
wb_notrcvd_1 = openpyxl.Workbook()
ws_notrcvd_1 = wb_notrcvd_1.active
for notrcvd_row in ws_agile_notrcvd.iter_rows(min_row=1,max_row=ws_agile_notrcvd.max_row+1,min_col=1,max_col=ws_agile_notrcvd.max_column+1,values_only=True):
#    print(sent_row)
    item_notrcvd = notrcvd_row[4]
    flag = False
#    print("agileorder_notrcvd item:" + str(item_notrcvd))
    for rcvd_row in ws_wacl_rcvd.iter_rows(min_row=2,max_row=ws_wacl_rcvd.max_row+1,min_col=1,max_col=ws_wacl_rcvd.max_column+1,values_only=True):
#        print(rcvd_row)
        item_rcvd = rcvd_row[4]
#        print("waclorder_rcvd item:" + str(item_rcvd))
        if item_notrcvd == item_rcvd:
            flag = True
            break;
    if flag == False:
        ws_notrcvd_1.append(notrcvd_row)

wb_notrcvd_1.save(output_excel_path+'Order_notrcvd_1.xlsx')

# Input:  Wacl_notrcvd.xlsx: Order NOT received from Wacl
# Input:  INV2308142-INS-5595.xlsx: Order recived from Agile
# Output: Order_notrcvd_2.xlsx: Order NOT received from  Agile and Wacl
wb_notrcvd_2 = openpyxl.Workbook()
ws_notrcvd_2 = wb_notrcvd_2.active
for sent_row in ws_wacl_notrcvd.iter_rows(min_row=1,max_row=ws_wacl_notrcvd.max_row+1,min_col=1,max_col=ws_wacl_notrcvd.max_column+1,values_only=True):
#    print(sent_row)
    item_notrcvd = sent_row[4]
    flag = False
#    print(item_sent)
    for rcvd_row in ws_agile_rcvd.iter_rows(min_row=2,max_row=ws_agile_rcvd.max_row+1,min_col=1,max_col=ws_agile_rcvd.max_column+1,values_only=True):
#        print(rcvd_row)
        item_rcvd = rcvd_row[6]
#        print(item_rcvd)
        if item_notrcvd == item_rcvd:
            flag = True
            break
    if flag == False:
        ws_notrcvd_2.append(sent_row)

wb_notrcvd_2.save(output_excel_path+'Order_notrcvd_2.xlsx')

# Input:  Order_notrcvd_1.xlsx: Order NOT received from  Agile and Wacl
# Input:  Order_notrcvd_2.xlsx: Order NOT received from  Agile and Wacl
# Output: Final_notrcvd.xlsx: Meged final Order NOT received from  Agile and Wacl
wb_fnotrcvd = openpyxl.load_workbook(output_excel_path+"Order_notrcvd_2.xlsx")
ws_fnotrcvd = wb_fnotrcvd.active

for sent_row in ws_notrcvd_1.iter_rows(min_row=1,max_row=ws_notrcvd_1.max_row+1,min_col=1,max_col=ws_notrcvd_1.max_column+1,values_only=True):
#    print(sent_row)
    item_notrcvd = sent_row[4]
    flag = False
#    print(item_sent)
    for rcvd_row in ws_notrcvd_2.iter_rows(min_row=2,max_row=ws_notrcvd_2.max_row+1,min_col=1,max_col=ws_notrcvd_2.max_column+1,values_only=True):
#        print(rcvd_row)
        item_rcvd = rcvd_row[4]
#        print(item_rcvd)
        if item_notrcvd == item_rcvd:
            flag = True
            break;
    if flag == False:
        ws_fnotrcvd.append(sent_row)

wb_fnotrcvd.save(output_excel_path+'Final_notrcvd.xlsx')
'''