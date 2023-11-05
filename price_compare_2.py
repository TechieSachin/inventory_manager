import openpyxl
import glob
import os
import tabula
import pandas as pd
from openpyxl.styles import Font
from openpyxl import Workbook
import csv

input_agile_path = "input/excel/agile/"
output_excel_path = "output/excel/"
csvname = "output.csv"
xlsxname = "output.xlsx"
# os.remove(output_excel_path+"Agile_fullrcvd.xlsx")  
# os.remove(output_excel_path+"Agile_fullrcvd_upc.xlsx")
# os.remove(output_excel_path+"*.xlsx")  

# save them in a folder
if not os.path.isdir(input_agile_path):
    os.mkdir(input_agile_path)
if not os.path.isdir(output_excel_path):
    os.mkdir(output_excel_path)
#else:
#    os.remove(output_excel_path+"Agile_fullrcvd.xlsx")  
#    os.remove(output_excel_path+"Agile_fullrcvd_upc.xlsx")

def remove_files_in_dir(directory_path):
    # List all files in the directory
    file_list = os.listdir(directory_path)

    # Iterate through the files and delete .xlsx files
    for file in file_list:
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory_path, file)
            try:
                os.remove(file_path)
                print(f"Deleted: {file_path}")
            except OSError as e:
                print(f"Error deleting {file_path}: {e}")

remove_files_in_dir(output_excel_path)

def pdf_to_csv(pdf_file,csvname):
    # Convert into Excel File
    tabula.convert_into(pdf_file, input_agile_path+csvname, output_format="csv", pages='all',stream=True)

def csv_to_excel(csvname,xlsxname):
    wb = Workbook()
    ws = wb.active
    with open(input_agile_path+csvname, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(input_agile_path+xlsxname)

# Function to convert a dollar amount string to a number
def dollar_string_to_number(dollar_string):
    if isinstance(dollar_string, str):
        try:
            # Remove dollar sign and commas, then convert to float
            return float(dollar_string.replace('$', '').replace(',', ''))
        except ValueError:
            return dollar_string  # Return the original value if conversion fails
    return dollar_string  # Return the original value if it's not a string

def fix_excel(fix_xl_file):
    upc_list_fix = []
    upc_dict_fix = {}
    upc_flag_fix = 0
    wb_fix_excel = openpyxl.Workbook()
    ws_fix_excel = wb_fix_excel.active
    ws_fix_excel.append(["Product", "Quantity", "Unit Price","Taxable","Total Price"])
    wb_xl_to_fix = openpyxl.load_workbook(fix_xl_file)
    ws_xl_to_fix = wb_xl_to_fix.active
    for rcvd_row in ws_xl_to_fix.iter_rows(min_row=1,max_row=ws_xl_to_fix.max_row,min_col=1,max_col=ws_xl_to_fix.max_column+1,values_only=True):
    #    ws_agile_rcvd.delete_rows(0, 1)
    #    wb_agile_rcvd.save(file_list[0])
#       print(rcvd_row)
        item0 = rcvd_row[0]
        item2 = rcvd_row[2]
    #    print(item0,item2)
        if item0 == "Fuel Surcharge": #Skip fuel
            continue
        elif ((item2 == None) and (upc_flag_fix == 1) and item0.isnumeric()):
            upc_flag_fix = 0
            upc_list_fix.append(item0)
            upc_dict_fix[item_desc] = item0
        elif ((item2 != None) and (upc_flag_fix == 1)):
            print("xlfix: Update",file,"csv with UPC for",item_desc)   
            upc_list_fix.append(None)
            ws_fix_excel.append(rcvd_row)
        elif (upc_flag_fix ==0):
            upc_flag_fix = 1
            item_desc= item0
            ws_fix_excel.append(rcvd_row)
    fix_xl_name=os.path.splitext(os.path.basename(fix_xl_file))[0]
    wb_fix_excel.save(output_excel_path + fix_xl_name+".xlsx")
    df = pd.read_excel(output_excel_path + fix_xl_name+".xlsx")
    # Apply the function to the 'Amount' column
    df['Unit Price'] = df['Unit Price'].apply(dollar_string_to_number)

    # Save the modified DataFrame back to the same Excel file
    df.to_excel(output_excel_path + fix_xl_name+".xlsx", index=False, engine='openpyxl')
#    wb_fix_excel = openpyxl.load_workbook(output_excel_path + fix_xl_name+".xlsx")
#    ws_fix_excel = wb_fix_excel.active
#    ws_fix_excel.delete_rows(0, 1)
#    wb_fix_excel.save(output_excel_path + fix_xl_name+".xlsx")


def price_compare(file1,file2):
    # print(file1)
    # print(file2)
    # Load the data from the first Excel file
    df1 = pd.read_excel(file1)
    #print(df1)
    # Load the data from the second Excel file
    df2 = pd.read_excel(file2)

    # Merge the two dataframes on the item names (assuming the item names are in a column called 'Item')
    merged_data = df1.merge(df2, on='Product', suffixes=('_file1', '_file2'))

    # Compare the item prices and create a new column for the price difference
    merged_data['Price_Difference'] = merged_data['Unit Price_file1'] - merged_data['Unit Price_file2']

    # You can now view or export the results
    #print(merged_data)
    file1_name=os.path.splitext(os.path.basename(file1))[0]
    file2_name=os.path.splitext(os.path.basename(file2))[0]
    print(file1_name)
    print(file2_name)
    # Export the results to a new Excel file
    merged_data.to_excel(f'{output_excel_path}/compare_{file1_name}_{file2_name}.xlsx', index=False)
    
    # Read the existing XLSX file
    df = pd.read_excel(f'{output_excel_path}/compare_{file1_name}_{file2_name}.xlsx')

    # Keep only the desired columns
    columns_to_keep = ['Product', 'Price_Difference', 'Unit Price_file1', 'Unit Price_file2']
    df = df[columns_to_keep]

    # Write the updated DataFrame back to the XLSX file
    df.to_excel(f'{output_excel_path}/compare_{file1_name}_{file2_name}.xlsx', index=False, engine='openpyxl')

    print("Columns have been filtered and saved back to " f'{output_excel_path}/compare_{file1_name}_{file2_name}.xlsx')


pdf_list = sorted(glob.glob("input/pdf" + "/*.pdf"))
print(pdf_list)
#pdf_name=os.path.splitext(os.path.basename(pdf_list[0]))[0]
#print(pdf_name)

for pdf_file in pdf_list:
    pdf_name=os.path.splitext(os.path.basename(pdf_file))[0]
#    print(pdf_name)
    pdf_to_csv(pdf_file,f"{pdf_name}.csv")
    csv_to_excel(f"{pdf_name}.csv",f"{pdf_name}.xlsx")

file_list = sorted(glob.glob(input_agile_path + "/*.xlsx"))
#print(file_list)
upc_list = []
upc_dict = {}
upc_flag = 0
wb_agile_rcvd_full = openpyxl.Workbook()
ws_agile_rcvd_full = wb_agile_rcvd_full.active
for file in file_list:
    fix_excel(file)
    wb_agile_rcvd = openpyxl.load_workbook(file)
    ws_agile_rcvd = wb_agile_rcvd.active
    for rcvd_row in ws_agile_rcvd.iter_rows(min_row=1,max_row=ws_agile_rcvd.max_row,min_col=1,max_col=ws_agile_rcvd.max_column+1,values_only=True):
    #    ws_agile_rcvd.delete_rows(0, 1)
    #    wb_agile_rcvd.save(file_list[0])
 #       print(rcvd_row)
        item0 = rcvd_row[0]
        item2 = rcvd_row[2]
    #    print(item0,item2)
        if item0 == "Fuel Surcharge": #Skip fuel
            continue
        elif ((item2 == None) and (upc_flag == 1) and item0.isnumeric()):
            upc_flag = 0
            upc_list.append(item0)
            upc_dict[item_desc] = item0
        elif ((item2 != None) and (upc_flag == 1)):
             print("Update",file,"csv with UPC for",item_desc)   
             upc_list.append(None)
             ws_agile_rcvd_full.append(rcvd_row)
        elif (upc_flag ==0):
            upc_flag = 1
            item_desc= item0
            ws_agile_rcvd_full.append(rcvd_row)

file_out_list = sorted(glob.glob(output_excel_path + "/*.xlsx"))
idxx = 1
for file in file_out_list:
    if(idxx < len(file_out_list)):
        price_compare(file,file_out_list[idxx])
        idxx=idxx+1

wb_agile_rcvd_full.save(output_excel_path + 'Agile_fullrcvd.xlsx')
#print(upc_list)
#print(len(upc_list))
#print(upc_dict)


wb_agile_rcvd_full = openpyxl.load_workbook(output_excel_path + 'Agile_fullrcvd.xlsx')
ws_agile_rcvd_full = wb_agile_rcvd_full.active

wb_agile_rcvd_full_upc = openpyxl.Workbook()
ws_agile_rcvd_full_upc = wb_agile_rcvd_full_upc.active
i=0
rcvd_row_upc = list()
ws_agile_rcvd_full_upc.append(["Product", "Quantity", "Unit Price","Taxable","Total Price", "SKU"])
for rows in ws_agile_rcvd_full_upc.iter_rows(min_row=1, max_row=1, min_col=None):
   for cell in rows:
     cell.font = Font(bold=True)

# ws_agile_rcvd_full_upc['F1'].font = Font(bold=True)
for rcvd_row in ws_agile_rcvd_full.iter_rows(min_row=1,max_row=ws_agile_rcvd_full.max_row,min_col=1,max_col=ws_agile_rcvd_full.max_column+1,values_only=True):
 #   print(i,rcvd_row)
    rcvd_row_upc = list(rcvd_row)
    rcvd_row_upc[5]=upc_list[i]
    if(upc_list[i] == None):
        print("UPC 'None' for",rcvd_row_upc[0], "Please enter manually!") 
    # if  rcvd_row_upc[0] in upc_dict:
    #     print("Dict", rcvd_row_upc[0], upc_dict[rcvd_row_upc[0]])
    # else:
    #     print("Dict: UPC 'None' for",rcvd_row_upc[0], "Please enter manually!") 

 #   rcvd_row_upc_list.append()
    i = i + 1
 #   print(rcvd_row_upc)
    ws_agile_rcvd_full_upc.append(rcvd_row_upc)
wb_agile_rcvd_full_upc.save(output_excel_path + 'Agile_fullrcvd_upc.xlsx')
