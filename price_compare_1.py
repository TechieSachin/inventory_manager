import openpyxl
import glob
import os
import tabula
import pandas as pd
from openpyxl.styles import Font

input_agile_path = "input/excel/agile/"
output_excel_path = "output/excel/"

# save them in a folder
if not os.path.isdir(input_agile_path):
    os.mkdir(input_agile_path)
if not os.path.isdir(output_excel_path):
    os.mkdir(output_excel_path)
#else:
#    os.remove(output_excel_path+"Agile_fullrcvd.xlsx")  
#    os.remove(output_excel_path+"Agile_fullrcvd_upc.xlsx")      
'''
pdf_list = sorted(glob.glob("input/pdf" + "/*.pdf"))
print(pdf_list)

#    print(os.path.splitext(os.path.basename(pdf_file))[0])
pdf_name=os.path.splitext(os.path.basename(pdf_list[0]))[0]

# read PDF file
#tables = tabula.read_pdf("input/pdf/Invoice160.pdf", pages="all")
tables = tabula.read_pdf(pdf_list[0], pages="all")

# iterate over extracted tables and export as excel individually
for i, table in enumerate(tables, start=1):
    table.to_excel(os.path.join(input_agile_path, f"excel_{pdf_name}_{i}.xlsx"), index=False)
'''
# excel files in the path
file_list = sorted(glob.glob(input_agile_path + "/*.xlsx"))
#print(file_list)

upc_list = []
upc_dict = {}
wb_agile_rcvd_full = openpyxl.Workbook()
ws_agile_rcvd_full = wb_agile_rcvd_full.active
for file in file_list:
    wb_agile_rcvd = openpyxl.load_workbook(file)
    ws_agile_rcvd = wb_agile_rcvd.active
    #status = ws_agile_rcvd['A1'].value
    #print(status)
    for rcvd_row in ws_agile_rcvd.iter_rows(min_row=1,max_row=ws_agile_rcvd.max_row,min_col=1,max_col=ws_agile_rcvd.max_column+1,values_only=True):
    #    ws_agile_rcvd.delete_rows(0, 1)
    #    wb_agile_rcvd.save(file_list[0])
        print(rcvd_row)
        item0 = rcvd_row[0]
    #    print(item0)
        item2 = rcvd_row[2]
    #    print(item2)
        if item0 == "Fuel Surcharge": #Skip fuel
            continue
        elif (item2 == None):
            upc_list.append(item0)
            upc_dict[item_desc] = item0
        else:
            item_desc= item0
            ws_agile_rcvd_full.append(rcvd_row)
wb_agile_rcvd_full.save(output_excel_path + 'Agile_fullrcvd.xlsx')
print(upc_list)
print(len(upc_list))
#print(upc_dict)

wb_agile_rcvd_full = openpyxl.load_workbook(output_excel_path + 'Agile_fullrcvd.xlsx')
#wb_agileorder_rcvd = openpyxl.load_workbook('input/excel/agile/INV2308142-INS-5595.xlsx')
ws_agile_rcvd_full = wb_agile_rcvd_full.active

wb_agile_rcvd_full_upc = openpyxl.Workbook()
ws_agile_rcvd_full_upc = wb_agile_rcvd_full_upc.active
i=0
rcvd_row_upc = list()
ws_agile_rcvd_full_upc.append(["Product", "Quantity", "Unit Price","Taxable","Total Price", "SKU"])
for rows in ws_agile_rcvd_full_upc.iter_rows(min_row=1, max_row=1, min_col=None):
   for cell in rows:
     cell.font = Font(bold=True)

# ws_agile_rcvd_full_upc['F1'].font = Font(bold=True)
for rcvd_row in ws_agile_rcvd_full.iter_rows(min_row=1,max_row=ws_agile_rcvd_full.max_row,min_col=1,max_col=ws_agile_rcvd_full.max_column+1,values_only=True):
    #rcvd_row_upc = rcvd_row
    print(i,rcvd_row)
 #   print(upc_list[i])
  #  print(list(rcvd_row) + list(upc_list[i]))
    rcvd_row_upc = list(rcvd_row)
    rcvd_row_upc[5]=upc_list[i]
    if(upc_list[i] == None):
        print("UPC 'None' for",rcvd_row_upc[0], "Please enter manually!") 
    # if  rcvd_row_upc[0] in upc_dict:
    #     print("Dict", rcvd_row_upc[0], upc_dict[rcvd_row_upc[0]])
    # else:
    #     print("Dict: UPC 'None' for",rcvd_row_upc[0], "Please enter manually!") 

 #   rcvd_row_upc_list.append()
    i = i + 1
 #   print(rcvd_row_upc)
    ws_agile_rcvd_full_upc.append(rcvd_row_upc)
wb_agile_rcvd_full_upc.save(output_excel_path + 'Agile_fullrcvd_upc.xlsx')

'''
excel_pd_list = []
for file in file_list:
    excel_pd_list.append(pd.read_excel(file))

# create a new dataframe to store the 
# merged excel file.
agile_merged_df = pd.DataFrame()
 
for excel_file in excel_pd_list:
    # appends the data into the excl_merged 
    # dataframe.
    agile_merged_df = agile_merged_df._append(excel_file, ignore_index=True)
 
# exports the dataframe into excel file with
# specified name.
agile_merged_df.to_excel(output_excel_path + 'agile_rcvd_merged.xlsx', index=False)
'''